"""
High School Management System API

A super simple FastAPI application that allows students to view and sign up
for extracurricular activities at Mergington High School.
"""

from fastapi import FastAPI, HTTPException
from fastapi.staticfiles import StaticFiles
from fastapi.responses import RedirectResponse
import os
from pathlib import Path
from openpyxl import load_workbook

app = FastAPI(title="Mergington High School API",
              description="API for viewing and signing up for extracurricular activities")

# Mount the static files directory
current_dir = Path(__file__).parent
app.mount("/static", StaticFiles(directory=os.path.join(Path(__file__).parent,
          "static")), name="static")

# In-memory activity database
default_activities = {
    "Chess Club": {
        "description": "Learn strategies and compete in chess tournaments",
        "schedule": "Fridays, 3:30 PM - 5:00 PM",
        "max_participants": 12,
        "participants": ["michael@mergington.edu", "daniel@mergington.edu"]
    },
    "Programming Class": {
        "description": "Learn programming fundamentals and build software projects",
        "schedule": "Tuesdays and Thursdays, 3:30 PM - 4:30 PM",
        "max_participants": 20,
        "participants": ["emma@mergington.edu", "sophia@mergington.edu"]
    },
    "Gym Class": {
        "description": "Physical education and sports activities",
        "schedule": "Mondays, Wednesdays, Fridays, 2:00 PM - 3:00 PM",
        "max_participants": 30,
        "participants": ["john@mergington.edu", "olivia@mergington.edu"]
    }
}

excel_path = current_dir / "data" / "activities.xlsx"


def load_activities_from_excel(path: Path):
    if not path.exists():
        return default_activities

    workbook = load_workbook(path, data_only=True)
    sheet = workbook.active
    loaded_activities = {}

    for row in sheet.iter_rows(min_row=2, values_only=True):
        name, description, schedule, max_participants, participants = row
        if not name:
            continue

        participants_list = []
        if participants:
            participants_list = [email.strip() for email in str(participants).split(",") if email.strip()]

        loaded_activities[str(name)] = {
            "description": str(description or ""),
            "schedule": str(schedule or ""),
            "max_participants": int(max_participants or 0),
            "participants": participants_list,
        }

    return loaded_activities or default_activities


activities = load_activities_from_excel(excel_path)


@app.get("/")
def root():
    return RedirectResponse(url="/static/index.html")


@app.get("/activities")
def get_activities():
    return activities


@app.get("/dashboard")
def get_dashboard():
    total_activities = len(activities)
    total_participants = sum(len(activity["participants"]) for activity in activities.values())
    total_capacity = sum(activity["max_participants"] for activity in activities.values())

    return {
        "total_activities": total_activities,
        "total_participants": total_participants,
        "total_capacity": total_capacity,
        "available_spots": total_capacity - total_participants,
    }


@app.post("/activities/{activity_name}/signup")
def signup_for_activity(activity_name: str, email: str):
    """Sign up a student for an activity"""
    # Validate activity exists
    if activity_name not in activities:
        raise HTTPException(status_code=404, detail="Activity not found")

    # Get the specific activity
    activity = activities[activity_name]

    # Add student
    activity["participants"].append(email)
    return {"message": f"Signed up {email} for {activity_name}"}
